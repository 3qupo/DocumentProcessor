import os
import sys
import json
import ctypes
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import warnings
warnings.filterwarnings('ignore')

class MuzlotoScanner:
    """Сканер анкет Muzloto с сохранением в один Excel файл."""
    
    # Поля анкеты Muzloto в правильном порядке
    FIELD_NAMES = [
        "Дата заполнения",      # Когда обработана анкета
        "Файл анкеты",          # Имя файла скана
        "Дата визита",          # Дата: 18.12
        "Номер столика",        # Номер столика:
        "Место игры",           # Место игры:
        "Довольны посещением",  # Довольны ли вы посещением Музлого?
        "Понравился плейлист",  # Понравился ли вам плейлист?
        "Треки для добавления", # Какие треки вы бы добавили?
        "Понравилась локация",  # Понравилась ли вам локация?
        "Понравились кухня и бар", # Понравилась ли вам кухня и бар?
        "Устроил сервис",       # Устроил ли вас сервис, время подачи?
        "Понравился ведущий",   # Понравилась ли вам работа ведущего?
        "Количество посещений", # Сколько раз вы были на Музлого?
        "Оценка стоимости",     # Оцените стоимость игры за билет
        "Знают о заказе",       # Знаете ли вы, что Музлого можно заказать?
        "Источник информации",  # Откуда вы о нас узнали?
        "Цель посещения",       # Ради чего вы обычно ходите на подобные вечеринки?
        "Предложения по улучшению", # Что нам стоит улучшить?
        "Телефон",              # Номер телефона (если оставлен)
        "Статус обработки",     # Успешно/Ошибка
        "Время обработки (мс)", # Сколько времени заняло
        "Сырой текст",          # Первые 500 символов распознанного текста
        "Оператор",             # Кто обработал анкету
        "Комментарий"           # Дополнительные заметки
    ]
    
    def __init__(self, 
                 excel_file: str = "анкеты_muzloto.xlsx",
                 tessdata_path: Optional[str] = None):
        """
        Args:
            excel_file: Путь к ОБЩЕМУ файлу Excel
            tessdata_path: Путь к данным Tesseract
        """
        self.excel_file = Path(excel_file)
        self.tessdata_path = tessdata_path
        
        # Загружаем C++ библиотеку
        self.lib = self._load_core_library()
        self.scanner_ptr = None
        
        # Инициализация
        self._init_scanner()
        self._ensure_excel_file()
        
        # Статистика
        self.stats = {
            "total": 0,
            "success": 0,
            "failed": 0,
            "last_file": None
        }
        
        print(f"✓ Сканер Muzloto инициализирован")
        print(f"  Файл для сохранения: {self.excel_file}")
    
    def _load_core_library(self):
        """Загружает скомпилированную C++ библиотеку."""
        # Определяем путь к библиотеке в зависимости от ОС
        if sys.platform == "win32":
            lib_name = "muzloto_core.dll"
            lib_path = Path(__file__).parent.parent / "build" / lib_name
        elif sys.platform == "darwin":
            lib_name = "libmuzloto_core.dylib"
            lib_path = Path(__file__).parent.parent / "build" / lib_name
        else:
            lib_name = "libmuzloto_core.so"
            lib_path = Path(__file__).parent.parent / "build" / lib_name
        
        if not lib_path.exists():
            # Пробуем другие пути
            search_paths = [
                Path(__file__).parent.parent / "lib" / lib_name,
                Path(sys.prefix) / "lib" / lib_name,
                Path.cwd() / lib_name
            ]
            
            for path in search_paths:
                if path.exists():
                    lib_path = path
                    break
            else:
                raise FileNotFoundError(
                    f"Не найдена библиотека {lib_name}. "
                    f"Скомпилируйте C++ ядро сначала."
                )
        
        # Загрузка библиотеки
        if sys.platform == "win32":
            return ctypes.CDLL(str(lib_path))
        else:
            return ctypes.CDLL(str(lib_path), ctypes.RTLD_GLOBAL)
    
    def _init_scanner(self):
        """Инициализация C++ сканера."""
        # Определяем функции
        self.lib.muzloto_create.restype = ctypes.c_void_p
        self.lib.muzloto_create.argtypes = []
        
        self.lib.muzloto_destroy.argtypes = [ctypes.c_void_p]
        
        self.lib.muzloto_initialize.argtypes = [
            ctypes.c_void_p, ctypes.c_char_p
        ]
        self.lib.muzloto_initialize.restype = ctypes.c_int
        
        self.lib.muzloto_scan_image.argtypes = [
            ctypes.c_void_p, ctypes.c_char_p
        ]
        self.lib.muzloto_scan_image.restype = ctypes.c_char_p
        
        self.lib.muzloto_free_string.argtypes = [ctypes.c_char_p]
        
        # Создаем сканер
        self.scanner_ptr = self.lib.muzloto_create()
        
        # Инициализируем с данными Tesseract
        tessdata = None
        if self.tessdata_path:
            tessdata = self.tessdata_path.encode('utf-8')
        
        result = self.lib.muzloto_initialize(self.scanner_ptr, tessdata)
        if result != 1:
            raise RuntimeError("Не удалось инициализировать C++ сканер")
    
    def _ensure_excel_file(self):
        """Создает или проверяет Excel файл с правильными колонками."""
        if not self.excel_file.exists():
            print(f"Создаю новый файл для анкет: {self.excel_file}")
            
            # Создаем DataFrame с нужными колонками
            df = pd.DataFrame(columns=self.FIELD_NAMES)
            
            # Сохраняем
            with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Анкеты')
            
            # Форматируем файл
            self._format_excel_file()
            
        else:
            # Проверяем, что файл имеет правильные колонки
            try:
                df = pd.read_excel(self.excel_file, sheet_name=0)
                existing_columns = list(df.columns)
                
                # Если колонки не совпадают, добавляем недостающие
                missing_columns = [col for col in self.FIELD_NAMES 
                                 if col not in existing_columns]
                
                if missing_columns:
                    print(f"Добавляю недостающие колонки: {missing_columns}")
                    
                    for col in missing_columns:
                        df[col] = ""
                    
                    # Сохраняем с новыми колонками
                    with pd.ExcelWriter(self.excel_file, engine='openpyxl', 
                                      mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, index=False, sheet_name='Анкеты')
                    
                    self._format_excel_file()
                    
            except Exception as e:
                print(f"Ошибка проверки файла Excel: {e}")
                # Создаем заново
                self.excel_file.unlink(missing_ok=True)
                self._ensure_excel_file()
    
    def _format_excel_file(self):
        """Форматирует Excel файл для лучшего вида."""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Настраиваем ширину колонок
            column_widths = {
                "A": 15,   # Дата заполнения
                "B": 20,   # Файл анкеты
                "C": 12,   # Дата визита
                "D": 12,   # Номер столика
                "E": 20,   # Место игры
                "F": 20,   # Довольны посещением
                "G": 20,   # Понравился плейлист
                "H": 25,   # Треки для добавления
                "I": 18,   # Понравилась локация
                "J": 22,   # Понравились кухня и бар
                "K": 20,   # Устроил сервис
                "L": 18,   # Понравился ведущий
                "M": 20,   # Количество посещений
                "N": 25,   # Оценка стоимости
                "O": 25,   # Знают о заказе
                "P": 25,   # Источник информации
                "Q": 30,   # Цель посещения
                "R": 30,   # Предложения по улучшению
                "S": 18,   # Телефон
                "T": 15,   # Статус обработки
                "U": 18,   # Время обработки
                "V": 40,   # Сырой текст
                "W": 15,   # Оператор
                "X": 25,   # Комментарий
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Стиль для заголовков
            header_fill = PatternFill(start_color="366092", 
                                    end_color="366092", 
                                    fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Применяем стиль к заголовкам
            for col in range(1, len(self.FIELD_NAMES) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Границы для всей таблицы
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Применяем границы к существующим данным
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                  min_col=1, max_col=len(self.FIELD_NAMES)):
                for cell in row:
                    cell.border = thin_border
            
            # Автофильтр
            ws.auto_filter.ref = ws.dimensions
            
            # Замораживаем заголовки
            ws.freeze_panes = "A2"
            
            wb.save(self.excel_file)
            print(f"✓ Файл отформатирован: {self.excel_file}")
            
        except Exception as e:
            print(f"⚠ Не удалось отформатировать Excel: {e}")
    
    def process_anketa(self, 
                      image_path: str,
                      operator: str = "Система",
                      comment: str = "") -> Dict[str, Any]:
        """
        Обрабатывает одну анкету и добавляет в общий Excel файл.
        
        Args:
            image_path: Путь к изображению анкеты
            operator: Имя оператора/пользователя
            comment: Дополнительный комментарий
            
        Returns:
            Результат обработки
        """
        result = {
            "success": False,
            "message": "",
            "row_number": None,
            "excel_file": str(self.excel_file),
            "scan_data": None
        }
        
        self.stats["total"] += 1
        self.stats["last_file"] = image_path
        
        try:
            # Проверяем файл
            image_path_obj = Path(image_path)
            if not image_path_obj.exists():
                raise FileNotFoundError(f"Файл не найден: {image_path}")
            
            print(f"\n📄 Обработка: {image_path_obj.name}")
            
            # Вызываем C++ ядро для распознавания
            scan_start = datetime.now()
            
            image_path_bytes = str(image_path_obj).encode('utf-8')
            json_str_ptr = self.lib.muzloto_scan_image(
                self.scanner_ptr, image_path_bytes
            )
            
            if not json_str_ptr:
                raise RuntimeError("C++ сканер вернул пустой результат")
            
            # Парсим JSON результат
            json_str = ctypes.string_at(json_str_ptr).decode('utf-8')
            self.lib.muzloto_free_string(json_str_ptr)
            
            scan_data = json.loads(json_str)
            
            scan_time = (datetime.now() - scan_start).total_seconds() * 1000
            
            if not scan_data.get("success", False):
                error_msg = scan_data.get("error_message", "Неизвестная ошибка")
                raise RuntimeError(f"Ошибка сканирования: {error_msg}")
            
            # Подготавливаем данные для Excel
            excel_row = self._prepare_excel_row(
                scan_data=scan_data,
                image_path=image_path_obj,
                operator=operator,
                comment=comment,
                processing_time_ms=scan_time
            )
            
            # Добавляем в Excel
            row_num = self._append_to_excel(excel_row)
            
            # Обновляем статистику
            self.stats["success"] += 1
            result.update({
                "success": True,
                "message": f"Анкета добавлена в строку {row_num}",
                "row_number": row_num,
                "scan_data": scan_data
            })
            
            print(f"✓ Успешно! Строка: {row_num}")
            print(f"  Дата: {scan_data.get('date', '—')}")
            print(f"  Столик: {scan_data.get('table_number', '—')}")
            if scan_data.get('phone_number'):
                print(f"  Телефон: {scan_data.get('phone_number')}")
            
        except Exception as e:
            self.stats["failed"] += 1
            result["message"] = f"Ошибка: {str(e)}"
            print(f"✗ Ошибка: {e}")
            
            # Записываем ошибку в Excel
            error_row = self._create_error_row(
                image_path=image_path,
                error=str(e),
                operator=operator
            )
            self._append_to_excel(error_row)
        
        return result
    
    def _prepare_excel_row(self, scan_data: Dict, image_path: Path,
                          operator: str, comment: str, 
                          processing_time_ms: float) -> Dict[str, Any]:
        """Создает строку для Excel из данных сканирования."""
        raw_text = scan_data.get('raw_text', '')
        if len(raw_text) > 500:
            raw_text = raw_text[:500] + "..."
        
        return {
            "Дата заполнения": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "Файл анкеты": image_path.name,
            "Дата визита": scan_data.get('date', ''),
            "Номер столика": scan_data.get('table_number', ''),
            "Место игры": scan_data.get('location', ''),
            "Довольны посещением": scan_data.get('satisfaction', ''),
            "Понравился плейлист": scan_data.get('playlist_liked', ''),
            "Треки для добавления": scan_data.get('tracks_to_add', ''),
            "Понравилась локация": scan_data.get('location_liked', ''),
            "Понравились кухня и бар": scan_data.get('kitchen_liked', ''),
            "Устроил сервис": scan_data.get('service_ok', ''),
            "Понравился ведущий": scan_data.get('host_work', ''),
            "Количество посещений": scan_data.get('visits_count', ''),
            "Оценка стоимости": scan_data.get('ticket_price', ''),
            "Знают о заказе": scan_data.get('know_booking', ''),
            "Источник информации": scan_data.get('source_info', ''),
            "Цель посещения": scan_data.get('purpose', ''),
            "Предложения по улучшению": scan_data.get('improvements', ''),
            "Телефон": scan_data.get('phone_number', ''),
            "Статус обработки": "Успешно",
            "Время обработки (мс)": round(processing_time_ms, 1),
            "Сырой текст": raw_text,
            "Оператор": operator,
            "Комментарий": comment
        }
    
    def _create_error_row(self, image_path: str, error: str, 
                         operator: str) -> Dict[str, Any]:
        """Создает строку с ошибкой для Excel."""
        return {
            "Дата заполнения": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "Файл анкеты": Path(image_path).name if image_path else "",
            "Дата визита": "",
            "Номер столика": "",
            "Место игры": "",
            "Довольны посещением": "",
            "Понравился плейлист": "",
            "Треки для добавления": "",
            "Понравилась локация": "",
            "Понравились кухня и бар": "",
            "Устроил сервис": "",
            "Понравился ведущий": "",
            "Количество посещений": "",
            "Оценка стоимости": "",
            "Знают о заказе": "",
            "Источник информации": "",
            "Цель посещения": "",
            "Предложения по улучшению": "",
            "Телефон": "",
            "Статус обработки": f"Ошибка: {error[:50]}",
            "Время обработки (мс)": "",
            "Сырой текст": "",
            "Оператор": operator,
            "Комментарий": "Ошибка обработки"
        }
    
    def _append_to_excel(self, row_data: Dict[str, Any]) -> int:
        """Добавляет строку в Excel файл и возвращает номер строки."""
        try:
            # Загружаем существующие данные
            df = pd.read_excel(self.excel_file, sheet_name=0)
            
            # Создаем DataFrame из новой строки
            new_row_df = pd.DataFrame([row_data])
            
            # Объединяем
            df = pd.concat([df, new_row_df], ignore_index=True)
            
            # Сохраняем обратно в Excel
            with pd.ExcelWriter(self.excel_file, engine='openpyxl', 
                              mode='w') as writer:
                df.to_excel(writer, index=False, sheet_name='Анкеты')
            
            # Возвращаем номер строки (Excel считает с 1, + заголовок)
            return len(df) + 1
            
        except Exception as e:
            print(f"Ошибка при сохранении в Excel: {e}")
            
            # Пробуем альтернативный метод через openpyxl
            try:
                wb = load_workbook(self.excel_file)
                ws = wb.active
                
                # Находим первую пустую строку
                next_row = ws.max_row + 1
                
                # Заполняем ячейки
                for i, field_name in enumerate(self.FIELD_NAMES, 1):
                    value = row_data.get(field_name, "")
                    ws.cell(row=next_row, column=i, value=value)
                
                wb.save(self.excel_file)
                return next_row
                
            except Exception as e2:
                raise RuntimeError(f"Не удалось сохранить в Excel: {e2}")
    
    def process_folder(self, 
                      folder_path: str,
                      operator: str = "Система",
                      file_patterns: List[str] = None) -> Dict[str, Any]:
        """
        Обрабатывает все анкеты в папке.
        
        Args:
            folder_path: Путь к папке со сканами
            operator: Имя оператора
            file_patterns: Шаблоны файлов (по умолчанию: *.jpg, *.png, *.jpeg)
            
        Returns:
            Статистика обработки
        """
        folder = Path(folder_path)
        if not folder.exists():
            return {
                "success": False,
                "message": f"Папка не найдена: {folder_path}",
                "processed": 0
            }
        
        if file_patterns is None:
            file_patterns = ["*.jpg", "*.png", "*.jpeg", "*.tiff", "*.bmp"]
        
        # Находим все файлы
        files = []
        for pattern in file_patterns:
            files.extend(folder.glob(pattern))
        
        files = sorted(files)  # Сортируем по имени
        
        if not files:
            return {
                "success": False,
                "message": f"Не найдено файлов в папке: {folder_path}",
                "processed": 0
            }
        
        print(f"\n📁 Обработка папки: {folder_path}")
        print(f"Найдено файлов: {len(files)}")
        
        results = {
            "total": len(files),
            "success": 0,
            "failed": 0,
            "details": []
        }
        
        for i, file_path in enumerate(files, 1):
            print(f"\n[{i}/{len(files)}] Обработка: {file_path.name}")
            
            result = self.process_anketa(
                image_path=str(file_path),
                operator=operator,
                comment=f"Пакетная обработка #{i}"
            )
            
            if result["success"]:
                results["success"] += 1
            else:
                results["failed"] += 1
            
            results["details"].append({
                "file": file_path.name,
                "success": result["success"],
                "message": result["message"],
                "row": result.get("row_number")
            })
            
            # Небольшая пауза между обработкой файлов
            import time
            time.sleep(0.1)
        
        print(f"\n{'='*50}")
        print(f"✅ ОБРАБОТКА ЗАВЕРШЕНА")
        print(f"   Успешно: {results['success']}")
        print(f"   С ошибками: {results['failed']}")
        print(f"   Всего: {results['total']}")
        print(f"   Файл с результатами: {self.excel_file}")
        print(f"{'='*50}")
        
        return results
    
    def get_statistics(self) -> Dict[str, Any]:
        """Возвращает статистику обработки."""
        # Читаем Excel файл для дополнительной статистики
        try:
            df = pd.read_excel(self.excel_file, sheet_name=0)
            total_rows = len(df)
            success_rows = len(df[df['Статус обработки'] == 'Успешно'])
            
            # Статистика по датам
            if 'Дата заполнения' in df.columns:
                dates = df['Дата заполнения'].dropna().unique()
                date_stats = len(dates)
            else:
                date_stats = 0
            
            return {
                "excel_file": str(self.excel_file),
                "total_records": total_rows,
                "successful_records": success_rows,
                "processing_stats": self.stats,
                "unique_dates": date_stats,
                "last_processed": self.stats.get("last_file")
            }
            
        except Exception as e:
            return {
                "error": str(e),
                "processing_stats": self.stats
            }
    
    def __del__(self):
        """Очистка ресурсов при удалении объекта."""
        if hasattr(self, 'scanner_ptr') and self.scanner_ptr:
            self.lib.muzloto_destroy(self.scanner_ptr)